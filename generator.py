# generator.py — Версия 7.0 (ФИНАЛЬНАЯ)
#
# ПРАВИЛА стиля (строгие):
# 1. НИКОГДА не трогаем заливку (fill) ячеек в существующих колонках
# 2. Для НОВЫХ колонок: add_new_column копирует всё от левого соседа (включая fill)
# 3. При записи данных: меняем ТОЛЬКО значение и цвет шрифта
#    - Красный  если значение изменилось
#    - Чёрный   если новое/без изменений
# 4. ЕДИНСТВЕННОЕ исключение: синий FFD9E1F2 → жёлтый, только там где пишем данные

import re
import copy
import os
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

MONTH_SHORT = {
    1:'янв', 2:'фев', 3:'мар', 4:'апр',
    5:'май', 6:'июн', 7:'июл', 8:'авг',
    9:'сен', 10:'окт', 11:'ноя', 12:'дек'
}
MONTH_SHORT_TO_NUM = {v: k for k, v in MONTH_SHORT.items()}

# Синий цвет Excel в шаблонах — заменяем на жёлтый при записи данных
BLUE_FILL_RGB = 'FFD9E1F2'
YELLOW_FILL   = PatternFill(fill_type='solid', fgColor='FFFFFF99')


def month_label(year: int, month: int) -> str:
    return f"{MONTH_SHORT[month]}.{str(year)[-2:]}"


def header_val_to_row_key(val) -> int:
    if val is None:
        return None
    if hasattr(val, 'year'):
        return val.year * 100 + val.month
    s = str(val).strip()
    try:
        n = int(float(s))
        if n > 100000:
            return n
    except (ValueError, TypeError):
        pass
    m = re.search(r'([а-яё]{3})[\.\-](\d{2})$', s, re.IGNORECASE)
    if m:
        mon = MONTH_SHORT_TO_NUM.get(m.group(1).lower())
        if mon:
            return (2000 + int(m.group(2))) * 100 + mon
    return None


def get_relevant_cutoff() -> int:
    """Март 2026 → 202602. Апрель 2026 → 202603."""
    now = datetime.now()
    if now.month == 1:
        return (now.year - 1) * 100 + 12
    return now.year * 100 + (now.month - 1)


def detect_structure(sheet) -> dict:
    mapping_col = None
    for col in range(1, sheet.max_column + 1):
        for row in range(1, 8):
            val = sheet.cell(row=row, column=col).value
            if val and any(kw in str(val).lower() for kw in ['маппинг', 'данные для маппинга', 'mapping']):
                mapping_col = col
                break
        if mapping_col:
            break

    best_header_row, best_count = 3, 0
    scan_up_to = mapping_col if mapping_col else sheet.max_column
    for row in range(1, 8):
        count = sum(
            1 for col in range(1, scan_up_to)
            if header_val_to_row_key(sheet.cell(row=row, column=col).value) is not None
        )
        if count > best_count:
            best_count = count
            best_header_row = row

    if mapping_col is None:
        last = 1
        for col in range(1, sheet.max_column + 1):
            if header_val_to_row_key(sheet.cell(row=best_header_row, column=col).value) is not None:
                last = col
        mapping_col = last + 1

    last_date_col = 1
    for col in range(1, mapping_col):
        if header_val_to_row_key(sheet.cell(row=best_header_row, column=col).value) is not None:
            last_date_col = col

    return {
        'header_row': best_header_row,
        'mapping_col': mapping_col,
        'last_date_col': last_date_col,
    }


def get_existing_date_cols(sheet, header_row: int, mapping_col: int) -> dict:
    result = {}
    for col in range(1, mapping_col):
        rk = header_val_to_row_key(sheet.cell(row=header_row, column=col).value)
        if rk and rk > 100000:
            result[rk] = col
    return result


def _get_rgb(color) -> str | None:
    """Безопасно читает RGB. None если тема/indexed."""
    try:
        if color.type == 'rgb':
            return color.rgb
    except Exception:
        pass
    return None


def _is_blue(cell) -> bool:
    """Возвращает True если у ячейки синий фон FFD9E1F2."""
    try:
        rgb = _get_rgb(cell.fill.fgColor)
        return rgb == BLUE_FILL_RGB
    except Exception:
        return False


def _write_value_and_font(cell, value, is_changed: bool):
    """
    Записывает значение и меняет ТОЛЬКО цвет шрифта.
    Если фон синий — заменяем на жёлтый.
    ВСЁ ОСТАЛЬНОЕ НЕ ТРОГАЕМ.
    """
    cell.value = value

    # Исправляем синий фон → жёлтый (только в этом случае)
    if _is_blue(cell):
        try:
            cell.fill = copy.copy(YELLOW_FILL)
        except Exception:
            pass

    # Меняем только цвет шрифта
    font_color = 'FFFF0000' if is_changed else 'FF000000'
    f = cell.font
    try:
        cell.font = Font(
            name=f.name, size=f.size, bold=f.bold, italic=f.italic,
            underline=f.underline, strike=f.strike, vertAlign=f.vertAlign,
            color=font_color,
        )
    except Exception:
        pass


def _values_equal(a, b) -> bool:
    try:
        return abs(float(a) - float(b)) < 1e-9
    except (TypeError, ValueError):
        return str(a) == str(b)


def add_new_column(sheet, after_col: int, header_row: int, year: int, month: int) -> int:
    """
    Вставляет новую колонку после after_col.
    Копирует ВСЕ стили (включая fill) от левого соседа — без изменений.
    """
    new_col = after_col + 1
    sheet.insert_cols(new_col)

    for row in range(1, sheet.max_row + 1):
        src  = sheet.cell(row=row, column=after_col)
        dest = sheet.cell(row=row, column=new_col)
        if src.has_style:
            try:
                dest.font      = copy.copy(src.font)
                dest.fill      = copy.copy(src.fill)
                dest.border    = copy.copy(src.border)
                dest.alignment = copy.copy(src.alignment)
                dest.number_format = src.number_format
            except Exception:
                pass

    # Заголовок новой колонки
    sheet.cell(row=header_row, column=new_col).value = datetime(year, month, 1)

    # Ширина колонки
    src_ltr = get_column_letter(after_col)
    new_ltr = get_column_letter(new_col)
    if src_ltr in sheet.column_dimensions:
        sheet.column_dimensions[new_ltr].width = sheet.column_dimensions[src_ltr].width

    return new_col


def process_sector(sector_key: str, config: dict, json_data: list) -> dict:
    template_path = config['template']
    output_path   = config['output']

    if not os.path.exists(template_path):
        return {'success': False, 'output': output_path, 'updated': 0,
                'added_cols': 0, 'error': f'Шаблон не найден: {template_path}'}

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    cutoff_rk = get_relevant_cutoff()

    # Фильтруем данные по cutoff
    data_lookup = {}
    for item in json_data:
        rk = item.get('row_key')
        if rk is not None:
            rk_int = int(rk)
            if rk_int > 100000 and rk_int <= cutoff_rk:
                data_lookup[rk_int] = item

    if not data_lookup:
        return {'success': False, 'output': output_path, 'updated': 0,
                'added_cols': 0, 'error': f'Нет данных (cutoff={cutoff_rk})'}

    try:
        wb    = openpyxl.load_workbook(template_path)
        sheet = wb.active
    except Exception as e:
        return {'success': False, 'output': output_path, 'updated': 0,
                'added_cols': 0, 'error': str(e)}

    struct      = detect_structure(sheet)
    header_row  = struct['header_row']
    mapping_col = struct['mapping_col']

    print(f"  [{sector_key}] header_row={header_row}, "
          f"mapping_col={mapping_col}({get_column_letter(mapping_col)}), cutoff={cutoff_rk}")

    # ── 1. Добавляем недостающие колонки ─────────────────────────────────────
    existing_cols = get_existing_date_cols(sheet, header_row, mapping_col)
    added_cols    = 0

    for rk in sorted(data_lookup.keys()):
        if rk in existing_cols:
            continue
        year, month = rk // 100, rk % 100

        # Ставим новую колонку после последней колонки с меньшим rk
        after_col = struct['last_date_col']
        for existing_rk in sorted(existing_cols):
            if existing_rk < rk:
                after_col = existing_cols[existing_rk]

        new_col    = add_new_column(sheet, after_col, header_row, year, month)
        added_cols += 1
        print(f"  [{sector_key}] Добавлена колонка {get_column_letter(new_col)}: {month_label(year, month)}")

        # Обновляем индексы после insert
        updated_existing = {}
        for k, v in existing_cols.items():
            updated_existing[k] = v + 1 if v > after_col else v
        updated_existing[rk] = new_col
        existing_cols = updated_existing
        mapping_col  += 1
        if new_col > struct['last_date_col']:
            struct['last_date_col'] = new_col

    # ── 2. Активные колонки = те у которых есть данные ───────────────────────
    active_columns = {col: rk for rk, col in existing_cols.items() if rk in data_lookup}

    if not active_columns:
        wb.save(output_path)
        return {'success': True, 'output': output_path, 'updated': 0,
                'added_cols': added_cols, 'error': None}

    updated     = 0
    filled_rows = {col: set() for col in active_columns}

    # ── 3. Проход 1: пишем данные из DataTable ────────────────────────────────
    for row in range(header_row + 1, sheet.max_row + 1):
        mk_val = sheet.cell(row=row, column=mapping_col).value
        if not mk_val:
            continue
        mk = str(mk_val).strip()
        if mk.lower() == 'формула':
            continue

        for col, rk in active_columns.items():
            target   = sheet.cell(row=row, column=col)
            old_val  = target.value

            # Пропускаем ячейки с обычными формулами (не внешние ссылки)
            if old_val and str(old_val).startswith('='):
                if not (str(old_val).startswith("='[") or str(old_val).startswith('="[')):
                    continue

            row_data = data_lookup.get(rk, {})
            if mk not in row_data or row_data[mk] is None:
                continue

            new_val    = row_data[mk]
            is_changed = (old_val is not None) and not _values_equal(old_val, new_val)

            _write_value_and_font(target, new_val, is_changed)
            filled_rows[col].add(row)
            updated += 1

    # ── 4. Проход 2: формулы ──────────────────────────────────────────────────
    first_col = min(active_columns.keys())

    for row in range(header_row + 1, sheet.max_row + 1):
        mk_val = sheet.cell(row=row, column=mapping_col).value
        if not mk_val or str(mk_val).strip().lower() != 'формула':
            continue

        for col, rk in active_columns.items():
            target  = sheet.cell(row=row, column=col)
            source  = sheet.cell(row=row, column=col - 1)
            old_val = target.value

            # Строим формулу
            if col == first_col:
                new_val = old_val
                if new_val is None or not str(new_val).startswith('='):
                    sv = source.value
                    if sv and str(sv).startswith('='):
                        new_val = Translator(str(sv), origin=source.coordinate)\
                            .translate_formula(target.coordinate)
                    else:
                        new_val = sv
            else:
                ref = sheet.cell(row=row, column=first_col)
                rv  = ref.value
                if rv and str(rv).startswith('='):
                    new_val = Translator(str(rv), origin=ref.coordinate)\
                        .translate_formula(target.coordinate)
                else:
                    sv = source.value
                    if sv and str(sv).startswith('='):
                        new_val = Translator(str(sv), origin=source.coordinate)\
                            .translate_formula(target.coordinate)
                    else:
                        new_val = sv

            if not new_val or not str(new_val).startswith('='):
                continue

            # Проверяем что все строки-источники уже заполнены
            cl       = get_column_letter(col)
            ref_rows = set(int(r) for r in re.findall(rf'{cl}(\d+)', str(new_val)))
            if ref_rows and not ref_rows.issubset(filled_rows[col]):
                continue

            is_changed = (old_val is not None) and (str(old_val) != str(new_val))

            _write_value_and_font(target, new_val, is_changed)
            filled_rows[col].add(row)
            updated += 1

    try:
        wb.save(output_path)
    except Exception as e:
        return {'success': False, 'output': output_path, 'updated': updated,
                'added_cols': added_cols, 'error': str(e)}

    return {'success': True, 'output': output_path, 'updated': updated,
            'added_cols': added_cols, 'error': None}


def build_combined_report(sector_results: dict, output_path: str) -> dict:
    SHEET_ORDER = [
        ('external', 'Внешний'),
        ('fiscal',   'Фискальный'),
        ('real',     'Реальный'),
        ('monetary', 'Монетарный'),
        ('social',   'Социальный'),
    ]

    combined_wb = openpyxl.Workbook()
    combined_wb.remove(combined_wb.active)
    errors, sheets_added = [], 0

    for sector_key, sheet_name in SHEET_ORDER:
        result = sector_results.get(sector_key, {})
        if not result.get('success'):
            errors.append(f"{sheet_name}: {result.get('error', 'нет результата')}")
            continue

        src_file = result.get('output')
        if not src_file or not os.path.exists(src_file):
            errors.append(f'{sheet_name}: файл не найден ({src_file})')
            continue

        try:
            src_wb    = openpyxl.load_workbook(src_file)
            src_sheet = src_wb.active
            new_sheet = combined_wb.create_sheet(title=sheet_name)

            for row in src_sheet.iter_rows():
                for cell in row:
                    nc = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        try:
                            nc.font      = copy.copy(cell.font)
                            nc.fill      = copy.copy(cell.fill)
                            nc.border    = copy.copy(cell.border)
                            nc.alignment = copy.copy(cell.alignment)
                            nc.number_format = cell.number_format
                        except Exception:
                            pass

            for cl, cd in src_sheet.column_dimensions.items():
                new_sheet.column_dimensions[cl].width = cd.width
            for rn, rd in src_sheet.row_dimensions.items():
                new_sheet.row_dimensions[rn].height = rd.height
            for mr in src_sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(mr))

            sheets_added += 1
        except Exception as e:
            errors.append(f'{sheet_name}: {e}')

    if sheets_added == 0:
        return {'success': False, 'output': output_path, 'sheets': 0,
                'error': 'Ни один лист не добавлен'}

    try:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        combined_wb.save(output_path)
    except Exception as e:
        return {'success': False, 'output': output_path, 'sheets': sheets_added, 'error': str(e)}

    return {'success': True, 'output': output_path, 'sheets': sheets_added,
            'errors': errors or None, 'error': None}
